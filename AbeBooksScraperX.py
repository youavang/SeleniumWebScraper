from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support import expected_conditions as EC
from openpyxl.reader.excel import load_workbook
import traceback
import pandas as pd
import time
import os

pwd = os.getcwd()

class AbeBooksScraperX:
    """This scraper will extract the title, author ISBNs from x amount of books from the first page it searched on abebooks.com."""
    
    def __init__(self, data, driver):
        """Parameter initialization"""
        self.title = data
        self.driver = driver

    def search_book(self):
        """This function search for a specific book on Open Library."""

        # search for book
        search = self.driver.find_element_by_class_name('gnav-searchbox-input')
        search.clear()
        search.send_keys(self.data)
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.CLASS_NAME, "gnav-searchbox-input"))).send_keys(Keys.RETURN)

    def get_info(self):
        """This function gets information of the book."""
        for m in range (1,x+1):
            try:
                Title = driver.find_element_by_xpath('.//*[@id="book-'+ str(m) + '"' + ']/div[2]/div[1]/div[1]/h2/a/span').text
                Author = driver.find_element_by_xpath('.//*[@id="book-'+ str(m)+ '"' + ']/div[2]/div[1]/div[1]/p[1]/strong').text
                ISBN10 = driver.find_element_by_xpath('.//*[@id="book-'+ str(m) + '"' + ']/div[2]/div[1]/div[1]/p[3]/a/span[1]').text.split()[2]
                ISBN13 = driver.find_element_by_xpath('.//*[@id="book-'+ str(m) + '"' + ']/div[2]/div[1]/div[1]/p[3]/a/span[2]').text.split()[2]
                Published = driver.find_element_by_xpath('.//*[@id="book-'+ str(m) + '"' + ']/div[2]/div[1]/div[1]/p[2]/span[3]').text

                items = {
                    'Title': Title,
                    'Author': Author,
                    'ISBN10': ISBN10,
                    'ISBN13': ISBN13,
                    'Published':Published}

                books.append(items)

            except NoSuchElementException:
                print(traceback.format_exc())
                print('Does not exist in database.')
                pass

    def new_search(self):
        """This function will scroll to the search box and clear it for the next search."""
        search = self.driver.find_element_by_class_name('gnav-searchbox-input')
        search.clear()      
    
    def close_session(self):
        """This function closes the session"""

        print('Your session has ended!')
        self.driver.close()

    def apply(self):
        """Apply to book search."""

        self.search_book()
        self.get_info()
        time.sleep(1)
        self.new_search()

if __name__ == '__main__':

    # Get file and loop through list.
    wb = load_workbook('Books5c.xlsx')
    sheet = wb['Sheet1']
    rows = sheet.max_row
    cols = sheet.max_column
  
    # create a list to store our data
    books=[]

    # How many books per search you want to get if found?
    x = 10

    #open website
    driver = webdriver.Chrome(executable_path ='C:\\chromedriver.exe')
    driver.get("https://www.abebooks.com")
    driver.maximize_window()
    

    # this will loop the scraper through the entire list from file
    for row in range(2, rows+1):
        for col in range(1, cols+1):
            data = sheet.cell(row, col).value
            bot = AbeBooksScraperX(data, driver)
            bot.apply()
            
    df = pd.DataFrame(books)
    df.to_excel('ScrapedBooks5.xlsx', index= None)
    bot.close_session()    

driver.close()