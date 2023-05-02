from openpyxl.reader.excel import load_workbook
from openpyxl.workbook.workbook import Workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.common.by import By
from selenium.common.exceptions import ElementClickInterceptedException, NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
import pandas as pd
import traceback
import time
import re
import json
import os

pwd = os.getcwd()

wb = load_workbook('MissingBooks3.xlsx')
sheet = wb['Sheet1']
rows = sheet.max_row
cols = sheet.max_column

data = []

driver = webdriver.Chrome(executable_path='C:\\chromedriver.exe')
driver.get("https://browse.nypl.org/iii/encore/search/?lang=eng")
driver.maximize_window()
time.sleep(2)

for row in range(2, rows+1):
    for col in range(1, cols+1):
        object = sheet.cell(row, col).value

        search = driver.find_element_by_css_selector('input#searchString')
        search.clear()
        search.send_keys(object)
        time.sleep(3)
        driver.find_element_by_css_selector('span.searchSubmit').click()

        books = driver.find_elements_by_xpath("//div[contains(@id, 'resultRecord')]")
        for book in books:

            try:
                driver.find_element_by_css_selector('span.title').click()
                time.sleep(5)
            except NoSuchElementException:
                pass

            try:
                title = driver.find_element_by_css_selector('div.dpBibTitle').text
                table = driver.find_element_by_css_selector('div#moreDetailsAnyComponent').text

                items = {
                    'Searched': object,
                    'Title': title,
                    'Description': table
                }

                data.append(items)

            except NoSuchElementException:
                pass
                
df = pd.DataFrame(data)
df.drop_duplicates(inplace=True)
df.to_excel('found4.xlsx', index = None)

driver.close()