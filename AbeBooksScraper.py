from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import NoSuchElementException, WebDriverException
from openpyxl.reader.excel import load_workbook
import traceback
import pandas as pd
import time
import os

pwd = os.getcwd()

class AbeBooksScraper:
    """This scraper will extract the title, author ISBNs from the first book it search on abebooks.com."""
    
    def __init__(self, data, driver):
        """Parameter initialization"""
        self.title = data
        self.driver = driver

    def search_book(self):
        """This function search for a specific book on Open Library."""

        # search for book
        search = self.driver.find_element_by_class_name('gnav-searchbox-input')
        search.clear()
        search.send_keys(self.title)
        self.driver.implicitly_wait(2)
        #time.sleep(2)
        search.send_keys(Keys.RETURN)

    def get_info(self):
        """This function gets information of the book."""
        self.driver.implicitly_wait(2)
        try:
            Title = self.driver.find_element_by_xpath('.//*[@id="book-1"]/div[2]/div[1]/div[1]/h2/a').text
            Author = self.driver.find_element_by_xpath('.//*[@id="book-1"]/div[2]/div[1]/div[1]/p[1]/strong').text
            ISBN10 = self.driver.find_element_by_xpath('.//*[@id="book-1"]/div[2]/div[1]/div[1]/p[3]/a/span[1]').text.split()[2]
            ISBN13 = self.driver.find_element_by_xpath('.//*[@id="book-1"]/div[2]/div[1]/div[1]/p[3]/a/span[2]').text.split()[2]
            Published = self.driver.find_element_by_xpath('.//*[@id="book-1"]/div[2]/div[1]/div[1]/p[2]/span[3]').text

            items = {
                'Title': Title,
                'Author': Author,
                'ISBN10': ISBN10,
                'ISBN13': ISBN13,
                'Published':Published}

            books.append(items)

        except NoSuchElementException:
            print(traceback.format_exc())
            print('Does not exist in database.')
            #self.refresh()
            #self.apply()
            pass
        except WebDriverException:
            #self.driver.refresh()
            #self.apply()
            pass

    def new_search(self):
        """This function will scroll to the search box and clear it for the next search."""
        search = self.driver.find_element_by_class_name('gnav-searchbox-input')
        search.clear()      
    
    def close_session(self):
        """This function closes the actual session"""

        print('Your session has ended!')
        self.driver.close()

    def apply(self):
        """Apply to book search."""

        self.driver.maximize_window()
        self.search_book()
        #time.sleep(2)
        self.get_info()
        time.sleep(1)
        self.new_search()

if __name__ == '__main__':

    # Get file and loop through list.
    wb = load_workbook('testing.xlsx')
    sheet = wb['Sheet1']
    rows = sheet.max_row
    cols = sheet.max_column
  
    # create a list to store our data
    books=[]

    #open website
    driver = webdriver.Chrome(executable_path ='C:\\chromedriver.exe')
    driver.get("https://www.abebooks.com")

    # this will loop the scraper through the entire list from file
    r = 0
    while r < rows:
        for row in range(2, rows+1):
            r += 1
            for col in range(1, cols+1):
                data = sheet.cell(row, col).value
                bot = AbeBooksScraper(data, driver)
                bot.apply()
        df = pd.DataFrame(books)
        df.to_excel('test_file6.xlsx', index= None)
        bot.close_session()
driver.refresh()
driver.close()