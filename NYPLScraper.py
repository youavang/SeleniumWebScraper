from openpyxl.reader.excel import load_workbook
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC, wait
from selenium.common.exceptions import  NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium import webdriver
import pandas as pd
import time
import os

pwd = os.getcwd()

class LACountyScraper:
    """This scraper will extract the ISBN, Title, Author, and Genre with a given data."""

    def __init__(self, data, driver):
        """Parameter initialization"""
        self.data = data
        self.driver = driver

    def search(self):
        """This function search for a sepecific book on LA County Library website."""

        time.sleep(2)
        search = driver.find_element_by_css_selector('input#searchString')
        search.clear()
        search.send_keys(object)
        time.sleep(4)
        driver.find_element_by_css_selector('span.searchSubmit').click()
        #WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'input#searchString'))).send_keys(Keys.RETURN)
        #WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'span.searchSubmit'))).click()

    def extract(self):
        """This function will get the ISBN, Title, Author, and Genre of a book."""

        try:   
            driver.find_element_by_css_selector('span.title').click()
            time.sleep(5)
        except NoSuchElementException:
            pass

        try:
            title = driver.find_element_by_css_selector('div.dpBibTitle').text
            table = driver.find_element_by_css_selector('div#moreDetailsAnyComponent').text

            items = {
                'Searched': object,
                'Title': title,
                'Description': table
            }

            books.append(items)
        except NoSuchElementException:
            no_item = {'Not Found': object}
            no_books.append(no_item)
            pass
            
        try:
            driver.find_element_by_css_selector('#nextRecordDisplayLinkComponent > span').click()
            time.sleep(3)
        except NoSuchElementException:
            pass

    def new_search(self):
        """This function will go back to the search box and clear it for the next search."""

        search = driver.find_element_by_css_selector('input#searchString')
        search.clear()
        self.driver.implicitly_wait(1)

    def close_session(self):
        """This funcition ends the session."""

        print('Your session has ended.')
        self.driver.close()

    def apply(self):
        """This function will invoke several functions to perform a search."""

        self.search()
        self.extract()
        self.new_search()

if __name__ == '__main__':

    wb = load_workbook('book2searching1.xlsx')
    sheet = wb['Sheet1']
    rows = sheet.max_row
    cols = sheet.max_column

    books = []
    no_books = []

    driver = webdriver.Chrome(executable_path ='C:\\chromedriver.exe')
    driver.get("https://browse.nypl.org/iii/encore/search/?lang=eng")
    driver.maximize_window()

    for row in range(1, rows+1):
        for col in range(1, cols+1):
            data = sheet.cell(row, col).value

            #data = isbn
            bot = LACountyScraper(data, driver)
            bot.apply()
    bot.close_session()

df = pd.DataFrame(books)
df.drop_duplicates(inplace=True)
df.to_excel('found2.xlsx', index = None)

dx = pd.DataFrame(no_books)
dx.drop_duplicates(inplace=True)
dx.to_excel('found_none2.xlsx', index = None)
