from openpyxl.reader.excel import load_workbook
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC, wait
from selenium.common.exceptions import  NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium import webdriver
import pandas as pd
import time
import os

pwd = os.getcwd()

class LACountyScraper:
    """This scraper will extract the ISBN, Title, Author, and Genre with a given data."""

    def __init__(self, data, driver):
        """Parameter initialization"""
        self.data = data
        self.driver = driver

    def search(self):
        """This function search for a sepecific book on LA County Library website."""

        time.sleep(2)
        search = self.driver.find_element_by_name('q')
        search.clear()
        search.send_keys(self.data)
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.NAME, 'q'))).send_keys(Keys.RETURN)

    def extract(self):
        """This function will get the ISBN, Title, Author, and Genre of a book."""

        try:
            # If one book is found, extract ISBN, Title, Author, and Genre.
            ISBN = self.driver.find_element_by_css_selector('div.displayElementText.text-p.ISBN').text
            Title = self.driver.find_element_by_css_selector('div.displayElementText.text-p.INITIAL_TITLE_SRCH').text
            Author = self.driver.find_element_by_css_selector('div.displayElementText.text-p.INITIAL_AUTHOR_SRCH>a').text
            Genre = self.driver.find_element_by_css_selector('div.displayElementText.text-p.SUBJECT_TERM').text

            items = {
                'Searched': data,
                'ISBN': ISBN,
                'Title': Title,
                'Author': Author,
                'Genre': Genre}

            books.append(items)
            
        except NoSuchElementException:
            # If there are multiple books listed, data from each book will be extracted.
            list = self.driver.find_elements_by_css_selector('div.results_cell')
            for item in list:
                self.driver.find_element_by_css_selector('img.results_img').click()
                self.driver.switch_to.active_element()
                #for x in range(len(list)):
                for x in item:
                #WebDriverWait(driver, 10).until(EC.visibility_of_element_located(By.CSS_SELECTOR, 'img.results_img')).click()
                    #self.driver.find_element_by_css_selector('img.results_img').click()
                    #self.driver.find_element_by_css_selector('img#syndeticsImg'+ str(x)).click()
                    #self.driver.switch_to_active_element()
                    time.sleep(2)
                    try:
                        ISBN = self.driver.find_element_by_css_selector('div.displayElementText.text-p.ISBN').text
                        Title = self.driver.find_element_by_css_selector('div.displayElementText.text-p.INITIAL_TITLE_SRCH').text
                        Author = self.driver.find_element_by_css_selector('div.displayElementText.text-p.INITIAL_AUTHOR_SRCH>a').text
                        try:
                            Genre = self.driver.find_element_by_css_selector('div.displayElementText.text-p.SUBJECT_TERM').text
                        except NoSuchElementException:
                            Genre = self.driver.find_element_by_css_selector('div.displayElementText.text-p.GENRE_TERM').text
                            try:
                                Genre = self.driver.find_element_by_css_selector('div.displayElementTable.SUBJECT').text
                            except NoSuchElementException:
                                Genre = self.driver.find_element_by_css_selector('div.asyncFieldSD_ITEM_STATUS').text

                        items = {
                        'Searched': data,
                        'ISBN': ISBN,
                        'Title': Title,
                        'Author': Author,
                        'Genre': Genre}

                    except NoSuchElementException:
                        ISBN = "none"
                        Title = self.driver.find_element_by_css_selector('div.displayElementText.text-p.TITLE').text
                        Author = self.driver.find_element_by_css_selector('div.displayElementText.text-p.AUTHOR').text
                        Genre = self.driver.find_element_by_css_selector('div.displayElementTable.SUBJECT').text
                    
                    except NoSuchElementException:
                        ISBN = "None"
                        Title = self.driver.find_element_by_css_selector('div.displayElementText.text-p.INITIAL_TITLE_SRCH').text
                        Author = "None"
                        Genre = self.driver.find_element_by_css_selector('div.displayElementText.text-p.GENRE_TERM').text

                        items = {
                        'Searched': data,
                        'ISBN': ISBN,
                        'Title': Title,
                        'Author': Author,
                        'Genre': Genre}
                        continue

                    books.append(items)
                    time.sleep(10)
                    self.driver.find_element_by_css_selector('a.nextArrowRight').click()
                    #WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'a.nextArrowRight'))).click()
                    time.sleep(5)
                    #self.driver.find_element_by_css_selector('button.ui-button.ui-corner-all.ui-widget.ui-button-icon-only.ui-dialog-titlebar-close').click()
                    self.driver.find_element_by_css_selector('button.ui-button.ui-corner-all.ui-widget').click()
                    time.sleep(2)
            
            try:
                log = {'Searched': data}

                no_books.append(log)     
            except NoSuchElementException:
                pass

    def new_search(self):
        """This function will go back to the search box and clear it for the next search."""

        search = self.driver.find_element_by_name('q')
        search.clear()
        self.driver.implicitly_wait(1)

    def close_session(self):
        """This funcition ends the session."""

        print('Your session has ended.')
        self.driver.close()

    def apply(self):
        """This function will invoke several functions to perform a search."""

        self.search()
        self.extract()
        self.new_search()

if __name__ == '__main__':

    wb = load_workbook('booksearching18.xlsx')
    sheet = wb['Sheet1']
    rows = sheet.max_row
    cols = sheet.max_column

    books = []
    no_books = []

    driver = webdriver.Chrome(executable_path ='C:\\chromedriver.exe')
    #driver = webdriver.Firefox(executable_path ='C:\\gecko.exe')
    driver.get("https://catalog.lacountylibrary.org/client/en_US/default")
    driver.maximize_window()

    for row in range(2, rows+1):
        for col in range(1, cols+1):
            data = sheet.cell(row, col).value

            #data = isbn
            bot = LACountyScraper(data, driver)
            bot.apply()
    bot.close_session()

df = pd.DataFrame(books)
df.drop_duplicates(inplace=True)
df.to_excel('book_found20.xlsx', index = None)

dx = pd.DataFrame(no_books)
dx.drop_duplicates(inplace=True)
dx.to_excel('book_none20.xlsx', index = None)
