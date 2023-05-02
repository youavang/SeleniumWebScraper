from openpyxl.reader.excel import load_workbook
from openpyxl.workbook.workbook import Workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.common.exceptions import  NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
import pandas as pd
import time
import os

pwd = os.getcwd()

# Open file and create variables.
wb = load_workbook('booksearcing17.xlsx')
sheet = wb['Sheet1']
rows = sheet.max_row
cols = sheet.max_column

#print(rows)
#print(cols)

books = []
no_books = []

driver = webdriver.Chrome(executable_path ='C:\\chromedriver.exe')
driver.get("https://catalog.lacountylibrary.org/client/en_US/default")
driver.maximize_window()

for row in range(2, rows+1):
    for col in range(1, cols+1):
        data = sheet.cell(row, col).value
        
        # Search for book via ISBN and only one book is listed.
        search = driver.find_element_by_name('q')
        search.clear()
        search.send_keys(data)
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.NAME, 'q'))).send_keys(Keys.RETURN)
        
        try:
            ISBN = driver.find_element_by_css_selector('div.displayElementText.text-p.ISBN').text
            Title = driver.find_element_by_css_selector('div.displayElementText.text-p.INITIAL_TITLE_SRCH').text
            Author = driver.find_element_by_css_selector('div.displayElementText.text-p.INITIAL_AUTHOR_SRCH>a').text
            Genre = driver.find_element_by_css_selector('div.displayElementText.text-p.SUBJECT_TERM').text

            items = {
                'ISBN': ISBN,
                'Title': Title,
                'Author': Author,
                'Genre': Genre}

            books.append(items)
            #pass
            
        except NoSuchElementException:
            # If there are multiple books listed, data from each book will be extracted.
            list = driver.find_elements_by_css_selector('div.results_cell')
            for item in list:
                #WebDriverWait(driver, 10).until(EC.visibility_of_element_located(By.CSS_SELECTOR, 'img.results_img')).click()
                driver.find_element_by_css_selector('img.results_img').click()
                driver.switch_to_active_element()
                time.sleep(2)
                try:
                    ISBN = driver.find_element_by_css_selector('div.displayElementText.text-p.ISBN').text
                    Title = driver.find_element_by_css_selector('div.displayElementText.text-p.INITIAL_TITLE_SRCH').text
                    Author = driver.find_element_by_css_selector('div.displayElementText.text-p.INITIAL_AUTHOR_SRCH>a').text
                    try:
                        Genre = driver.find_element_by_css_selector('div.displayElementText.text-p.SUBJECT_TERM').text
                    except NoSuchElementException:
                        Genre = driver.find_element_by_css_selector('div.displayElementText.text-p.GENRE_TERM').text
                        try:
                            Genre = driver.find_element_by_css_selector('div.displayElementTable.SUBJECT').text
                        except NoSuchElementException:
                            Genre = driver.find_element_by_css_selector('div.asyncFieldSD_ITEM_STATUS').text

                except NoSuchElementException:
                    ISBN = "none"
                    Title = driver.find_element_by_css_selector('div.displayElementText.text-p.TITLE').text
                    Author = driver.find_element_by_css_selector('div.displayElementText.text-p.AUTHOR').text
                    Genre = driver.find_element_by_css_selector('div.displayElementTable.SUBJECT').text

                    items = {
                    'Searched': data,
                    'ISBN': ISBN,
                    'Title': Title,
                    'Author': Author,
                    'Genre': Genre}
                
                try:
                    ISBN = "none"
                    Title = driver.find_element_by_css_selector('div.displayElementText.text-p.INITIAL_TITLE_SRCH').text
                    Author = "none"
                    Genre = driver.find_element_by_css_selector('div.displayElementText.text-p.GENRE_TERM').text
                
                except NoSuchElementException:
                    ISBN = "none"
                    Title = driver.find_element_by_css_selector('div.displayElementText.text-p.TITLE').text
                    Author ="none"
                    Genre = driver.find_element_by_css_selector('div.displayElementTable.SUBJECT').text

                    items = {
                    'Searched': data,
                    'ISBN': ISBN,
                    'Title': Title,
                    'Author': Author,
                    'Genre': Genre}

                books.append(items)
                driver.find_element_by_css_selector('button.ui-button.ui-corner-all.ui-widget.ui-button-icon-only.ui-dialog-titlebar-close').click()
                time.sleep(2)
            
            try:
                log = {'Searched': data}

                no_books.append(log)     
            except NoSuchElementException:
                pass

        search = driver.find_element_by_name('q')
        search.clear()
        driver.implicitly_wait(1)

    df = pd.DataFrame(books)
    df.to_excel('book_found19.xlsx', index=False)

    dx = pd.DataFrame(no_books)
    dx.to_excel('book_none19.xlsx', index= False)
driver.close()